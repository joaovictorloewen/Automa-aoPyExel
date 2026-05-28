import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from logger import logger
from config import ARQUIVO_RELATORIO

def gerar_relatorio(
    df,
    faturamento_total,
    ranking_vendedores,
    produtos_mais_vendidos
):

    logger.info("Gerando relatório")

    with pd.ExcelWriter(
        ARQUIVO_RELATORIO,
        engine="openpyxl"
    ) as writer:

        df.to_excel(
            writer,
            sheet_name="Vendas",
            index=False
        )

        ranking_vendedores.to_excel(
            writer,
            sheet_name="Ranking",
            index=False
        )

        produtos_mais_vendidos.to_excel(
            writer,
            sheet_name="Produtos",
            index=False
        )

        resumo = pd.DataFrame({
            "Indicador": ["Faturamento Total"],
            "Valor": [faturamento_total]
        })

        resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

def formatar_excel():

    logger.info("Formatando Excel")

    wb = load_workbook(ARQUIVO_RELATORIO)

    cor_header = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    for ws in wb.worksheets:

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = cor_header

        for column in ws.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 2

            ws.column_dimensions[
                column_letter
            ].width = adjusted_width

    wb.save(ARQUIVO_RELATORIO)